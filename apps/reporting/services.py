# apps/reporting/services.py
"""
Сервіси для генерації звітів та аналітики
"""

from django.db.models import Count, Q, F, Sum, Avg
from django.utils import timezone
from datetime import datetime, timedelta
from apps.personnel.models import Serviceman, Contract, ServiceHistoryEvent, Rank
from apps.staffing.models import Unit, Position, MilitarySpecialty
import pandas as pd
from typing import Dict, List, Any


class StaffingReportService:
    """Сервіс для звітів по укомплектованості"""

    @staticmethod
    def get_unit_staffing_report(unit_id: int) -> Dict[str, Any]:
        """
        Детальний звіт про укомплектованість підрозділу
        """
        unit = Unit.objects.get(pk=unit_id)

        # Отримуємо всі позиції в підрозділі та його підпорядкованих
        all_units = unit.get_descendants(include_self=True)
        positions = Position.objects.filter(unit__in=all_units)

        # Рахуємо статистику
        total_positions = positions.count()
        filled_positions = positions.filter(serviceman__isnull=False).count()
        vacant_positions = total_positions - filled_positions

        # Групування по категоріях
        by_category = positions.values('category').annotate(
            total=Count('id'),
            filled=Count('serviceman'),
        ).order_by('category')

        # Групування по ВОС
        by_specialty = positions.values('specialty__code', 'specialty__name').annotate(
            total=Count('id'),
            filled=Count('serviceman'),
        ).order_by('specialty__code')

        return {
            'unit': unit,
            'unit_name': unit.name,
            'date': timezone.now(),
            'summary': {
                'total_positions': total_positions,
                'filled_positions': filled_positions,
                'vacant_positions': vacant_positions,
                'staffing_percentage': round((filled_positions / total_positions * 100) if total_positions > 0 else 0,
                                             2),
            },
            'by_category': list(by_category),
            'by_specialty': list(by_specialty),
            'vacant_positions_list': positions.filter(serviceman__isnull=True).select_related('unit', 'specialty'),
        }

    @staticmethod
    def get_brigade_staffing_summary() -> List[Dict[str, Any]]:
        """
        Зведений звіт по всіх батальйонах бригади
        """
        battalions = Unit.objects.filter(name__contains='батальйон')
        summary = []

        for battalion in battalions:
            positions = Position.objects.filter(
                Q(unit=battalion) | Q(unit__in=battalion.get_descendants())
            )

            total = positions.count()
            filled = positions.filter(serviceman__isnull=False).count()

            summary.append({
                'battalion': battalion.name,
                'total_positions': total,
                'filled_positions': filled,
                'vacant_positions': total - filled,
                'percentage': round((filled / total * 100) if total > 0 else 0, 2),
            })

        return sorted(summary, key=lambda x: x['percentage'])


class PersonnelReportService:
    """Сервіс для звітів по особовому складу"""

    @staticmethod
    def get_personnel_statistics() -> Dict[str, Any]:
        """
        Загальна статистика по особовому складу
        """
        total_servicemen = Serviceman.objects.count()

        # По званнях
        by_rank = Serviceman.objects.values('rank__name').annotate(
            count=Count('id')
        ).order_by('rank__order')

        # По віку
        today = timezone.now().date()
        age_groups = {
            '18-25': 0,
            '26-30': 0,
            '31-35': 0,
            '36-40': 0,
            '41-45': 0,
            '46+': 0,
        }

        for serviceman in Serviceman.objects.all():
            age = (today - serviceman.date_of_birth).days // 365
            if age <= 25:
                age_groups['18-25'] += 1
            elif age <= 30:
                age_groups['26-30'] += 1
            elif age <= 35:
                age_groups['31-35'] += 1
            elif age <= 40:
                age_groups['36-40'] += 1
            elif age <= 45:
                age_groups['41-45'] += 1
            else:
                age_groups['46+'] += 1

        # По типу служби (контракт/мобілізація)
        contracts_ending_soon = Contract.objects.filter(
            end_date__lte=today + timedelta(days=90),
            end_date__gte=today
        ).count()

        return {
            'total_servicemen': total_servicemen,
            'by_rank': list(by_rank),
            'by_age': age_groups,
            'contracts_ending_soon': contracts_ending_soon,
            'average_age': Serviceman.objects.aggregate(
                avg_age=Avg(F('date_of_birth'))
            )['avg_age'],
        }

    @staticmethod
    def get_service_history_report(start_date: datetime, end_date: datetime) -> Dict[str, Any]:
        """
        Звіт по подіях в історії служби за період
        """
        events = ServiceHistoryEvent.objects.filter(
            event_date__gte=start_date,
            event_date__lte=end_date
        )

        by_type = events.values('event_type').annotate(
            count=Count('id')
        ).order_by('-count')

        return {
            'period': f"{start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}",
            'total_events': events.count(),
            'by_type': list(by_type),
            'events': events.select_related('serviceman', 'serviceman__rank').order_by('-event_date'),
        }


class ContractReportService:
    """Сервіс для звітів по контрактах"""

    @staticmethod
    def get_contracts_status() -> Dict[str, Any]:
        """
        Звіт по статусу контрактів
        """
        today = timezone.now().date()

        # Контракти, що закінчуються
        ending_30_days = Contract.objects.filter(
            end_date__gte=today,
            end_date__lte=today + timedelta(days=30)
        ).select_related('serviceman', 'serviceman__rank', 'serviceman__position')

        ending_90_days = Contract.objects.filter(
            end_date__gte=today + timedelta(days=31),
            end_date__lte=today + timedelta(days=90)
        ).select_related('serviceman', 'serviceman__rank', 'serviceman__position')

        expired = Contract.objects.filter(
            end_date__lt=today
        ).select_related('serviceman', 'serviceman__rank', 'serviceman__position')

        return {
            'date': today,
            'ending_30_days': {
                'count': ending_30_days.count(),
                'list': ending_30_days,
            },
            'ending_90_days': {
                'count': ending_90_days.count(),
                'list': ending_90_days,
            },
            'expired': {
                'count': expired.count(),
                'list': expired[:20],  # Показуємо тільки останні 20
            },
        }

    @staticmethod
    def get_contract_renewal_forecast() -> List[Dict[str, Any]]:
        """
        Прогноз по закінченню контрактів на наступні 12 місяців
        """
        today = timezone.now().date()
        forecast = []

        for month in range(12):
            start_date = today + timedelta(days=month * 30)
            end_date = today + timedelta(days=(month + 1) * 30)

            count = Contract.objects.filter(
                end_date__gte=start_date,
                end_date__lt=end_date
            ).count()

            forecast.append({
                'month': start_date.strftime('%B %Y'),
                'count': count,
            })

        return forecast


class ExportService:
    """Сервіс для експорту звітів"""

    @staticmethod
    def export_to_excel(report_data: Dict[str, Any], report_type: str) -> bytes:
        """
        Експорт звіту в Excel
        """
        import io
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = report_type

        # Заголовок
        ws['A1'] = f"Звіт: {report_type}"
        ws['A2'] = f"Дата: {timezone.now().strftime('%d.%m.%Y %H:%M')}"

        # Запис даних залежно від типу звіту
        row = 4
        if 'summary' in report_data:
            ws[f'A{row}'] = 'Загальна інформація'
            row += 1
            for key, value in report_data['summary'].items():
                ws[f'A{row}'] = key
                ws[f'B{row}'] = value
                row += 1

        # Зберігаємо в пам'яті
        virtual_workbook = io.BytesIO()
        wb.save(virtual_workbook)
        return virtual_workbook.getvalue()

    @staticmethod
    def export_to_pdf(report_data: Dict[str, Any], report_type: str) -> bytes:
        """
        Експорт звіту в PDF (потребує додаткових бібліотек)
        """
        # TODO: Імплементація PDF експорту
        pass